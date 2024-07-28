import datetime
import json
import os
from io import StringIO
from typing import Any, Callable

import pandas
from django.contrib.auth.models import AbstractUser
from django.db import models
from django.db.models.query import QuerySet
from django.http import HttpRequest, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from app import settings
from common import serizalize
from common.exception import ApiForbiddenError, ApiNotFoundError
from common.export import XlsxExportConfig
from common.models import Model
from common.response import ApiJsonResponse
from common.router import Router


def progress_get_query_params(request: HttpRequest):
    """_summary_
        make  page=1&size=10&browser__in[]=safari&browser__in[]=firefox
        to {'page': '1', 'size': '10', 'browser__in': ['safari', 'firefox']}
    Args:
        request (HttpRequest): _description_

    Returns:
        _type_: _description_
    """

    # {'size': "['10']", 'browser__in': ['safari', 'firefox', 'ie']
    list_params = request.GET.lists()
    print("list_params:", list_params)

    def replace_key(key):
        if key.endswith("[]"):
            return key[:-2]
        return key

    # list类型保留数组，其他的取第一个值
    return {
        **{
            replace_key(key): value[0] if not key.endswith("[]") else value
            for key, value in list_params
        }
    }


class Api:
    class Config:
        enable_create = True
        enable_update = True
        enable_delete = True

        def __init__(self, enable_create=True, enable_update=True, enable_delete=True):
            self.enable_create = enable_create
            self.enable_update = enable_update
            self.enable_delete = enable_delete

    model: Model | models.Model
    get_list_params: Callable[[HttpRequest], dict]
    get_create_params: Callable[[HttpRequest], dict]
    get_update_params: Callable[[HttpRequest], dict] | None
    get_export_params: Callable[[HttpRequest], dict]
    extra: dict[str, Callable[[models.Model], Any]]
    hidden: list[str]
    config: Config
    xlsx_config: XlsxExportConfig

    def __init__(
        self,
        model: models.base.ModelBase,
        get_list_params: Callable[[HttpRequest], dict] = lambda request: {
            **progress_get_query_params(request),
        },
        get_export_params: Callable[[HttpRequest], dict] = None,  # type: ignore
        get_create_params: Callable[[HttpRequest], dict] = lambda request: {
            **json.loads(request.body)
        },
        get_update_params: Callable[[HttpRequest], dict] | None = None,
        extra: dict[str, Callable[[models.Model], Any]] = {},
        config: Config = Config(),
        hidden: list[str] = [],
        xlsx_config: XlsxExportConfig = XlsxExportConfig([]),
    ):
        """__init__ method for Api

        Args:
            model (models.base.ModelBase): _description_ 数据模型
            get_list_params (_type_, optional): _description_. Defaults to lambdarequest:{ **request.GET.dict(), }.
            get_create_params (_type_, optional): _description_. Defaults to lambdarequest:{ **request.POST.dict() }.
            get_update_params (_type_, optional): _description_. Defaults to lambdarequest:{ **request.POST.dict() }.
            extra (dict[str, Callable[[models.Model], Any]], optional): _description_. Defaults to {}.
            config (Config, optional): _description_. Defaults to Config().
            hidden (list[str], optional): _description_. Defaults to [].
        """
        self.model = model  # type: ignore pylance 检测类型，数据模型总是 ModelBase，实际上是 Model，而且 ModelBase 也没有 orm 方法
        self.get_list_params = get_list_params
        self.get_create_params = get_create_params
        self.get_update_params = get_update_params
        self.get_export_params = get_export_params
        self.extra = extra
        self.config = config
        self.hidden = hidden
        self.xlsx_config = xlsx_config

    def get_hidden_fields(self):
        model_hidden_fields = getattr(self.model, "hidden_fields", [])
        return self.hidden + model_hidden_fields

    def get_protected_fields(self):
        model_protected_fields = getattr(self.model, "protected_fields", [])
        return model_protected_fields

    def get_xlsx_config(self):
        model_xlsx_config = getattr(self.model, "xlsx_config", XlsxExportConfig([]))
        assert isinstance(model_xlsx_config, XlsxExportConfig)
        return model_xlsx_config + self.xlsx_config

    def get_xlsx_fields(self):
        return self.get_xlsx_config().fields

    def extra_search_condition(self, request: HttpRequest):
        """多余的查询条件，比如从 header 获取 user_id，添加到查询条件中

        Args:
            request (HttpRequest): _description_

        Returns:
            _type_: _description_
        """
        return {}

    def serizalize(self, obj, **kwargs):
        return serizalize.serizalize(
            obj, extra=self.extra, hidden=self.get_hidden_fields(), **kwargs
        )

    def get(self, request: HttpRequest):
        id = request.GET.get("id")
        if not id:
            raise ApiNotFoundError("id is required")
        obj = self.model.objects.filter(**self.extra_search_condition(request)).get(
            pk=id
        )
        if obj:
            return ApiJsonResponse.success(self.serizalize(obj))
        raise ApiNotFoundError()

    def list_order_by(self, query: QuerySet, orders: list[str]):
        for order in orders:
            key, order = order.split("__")
            if order == "desc":
                query = query.order_by(f"-{key}")
            else:
                query = query.order_by(key)
        return query

    def modify_query(self, query: QuerySet):
        return query

    def list(self, request: HttpRequest):
        params = self.get_list_params(request)
        print("list params:", params)
        page = int(params.pop("page", 1))
        size = int(params.pop("size", 10))
        orders = str(params.pop("order_by", "id__desc")).split(",")

        query = self.model.objects.filter(**params).filter(
            **self.extra_search_condition(request)
        )
        query = self.list_order_by(query, orders)
        query = self.modify_query(query)

        # print sql
        print("list api:", query.query)

        total = query.count()
        data = query[(page - 1) * size : page * size]
        page_count = total // size + 1 if total % size > 0 else total // size
        return ApiJsonResponse.success(
            {
                "list": [self.serizalize(obj) for obj in data],
                "pageable": {
                    "total": total,
                    "page": page,
                    "size": size,
                    "page_count": page_count,
                },
            }
        )

    def create(self, request: HttpRequest):
        if not self.config.enable_create:
            raise ApiForbiddenError("Create is not allowed")
        params = self.get_create_params(request)
        protected_fields = self.get_protected_fields()
        for key in protected_fields:
            if key in params.copy():
                params.pop(key)
        obj = self.model.objects.create(**params)
        return ApiJsonResponse.success(self.serizalize(obj))

    def update(self, request: HttpRequest):
        if (
            not self.config.enable_update
            or self.model._meta.db_table in settings.LOCKING_MODIFY_TABLES
        ):
            raise ApiForbiddenError("Update is not allowed")
        params = (
            self.get_update_params(request)
            if self.get_update_params
            else self.get_create_params(request)
        )
        for key in self.get_protected_fields():
            if key in params.copy():
                params.pop(key)
        id = params.get("id")
        if not id:
            raise ApiNotFoundError("id is required")
        obj = self.model.objects.filter(**self.extra_search_condition(request)).get(
            pk=id
        )
        if not obj:
            raise ApiNotFoundError()
        for key, value in params.items():
            setattr(obj, key, value)
        obj.save()
        return ApiJsonResponse.success(self.serizalize(obj))

    def delete(self, request: HttpRequest):
        if (
            not self.config.enable_delete
            or self.model._meta.db_table in settings.LOCKING_DELETE_TABLES
        ):
            raise ApiForbiddenError("Delete is not allowed")
        params = {**json.loads(request.body)}
        ids = params.get("ids")
        id = params.get("id")

        if not ids and not id:
            raise ApiNotFoundError("ids/id is required")
        if id:
            ids = [id]
        query = self.model.objects.filter(
            **self.extra_search_condition(request)
        ).filter(pk__in=ids)
        for obj in query:
            obj.delete()
        return ApiJsonResponse.success("Delete Success")

    def export(self, request: HttpRequest):
        params = (
            self.get_export_params(request)
            if self.get_export_params
            else self.get_list_params(request)
        )
        orders = str(params.pop("order_by", "id__desc")).split(",")

        # 组织查询
        query = self.model.objects.filter(**params).filter(
            **self.extra_search_condition(request)
        )
        query = self.list_order_by(query, orders)
        query = self.modify_query(query)

        # print sql
        print("export api:", query.query)
        data = query
        json_data = [self.serizalize(obj, with_foreign_keys=False) for obj in data]
        # print("export data:",json_data)
        dataframe = pandas.read_json(StringIO(json.dumps(json_data)))
        file = self.get_export_filename()
        self.to_excel(dataframe, file)
        return HttpResponse(
            open(file, "rb"),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={self.model._meta.model_name}.xlsx"
            },
        )

    def get_export_filename(self):
        out_dir = os.path.join(settings.BASE_DIR, "temp/export")
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)
        date_str = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        file = os.path.join(out_dir, f"{self.model._meta.model_name}_{date_str}.xlsx")
        return file

    def is_unicode_chinese(self, value):
        for ch in value:
            if "\u4e00" <= ch <= "\u9fff":
                return True
        return False

    def to_excel(self, dataframe: pandas.DataFrame, file):
        wb = Workbook(write_only=False)
        ws: Worksheet = wb.active  # type: ignore
        assert ws, "ws is None"
        # print(dataframe)

        dataframe = dataframe.fillna("/")

        for field in self.get_xlsx_fields():
            if hasattr(dataframe, field.prop):
                dataframe[field.prop] = dataframe[field.prop].map(
                    lambda x: field.format(x)
                )

        # sort cols
        if len(self.get_xlsx_fields()) > 0:
            dataframe = dataframe.reindex(
                columns=[field.prop for field in self.get_xlsx_fields()]
            )

            sum_row = [None for _ in range(len(dataframe.columns))]
            sum_row[0] = "合计"  # type: ignore
            sub_dataframe = pandas.DataFrame([sum_row], columns=dataframe.columns)
            for field in self.get_xlsx_fields():
                if field.sum:
                    sub_dataframe[field.prop] = dataframe[field.prop].sum()
            dataframe = pandas.concat([dataframe, sub_dataframe])

        for r in dataframe_to_rows(dataframe, index=False, header=True):
            ws.append(r)

        # 自动宽度
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            length = min(50, length)
            width = 2.0 if self.is_unicode_chinese(column_cells[0].value) else 1.0
            ws.column_dimensions[column_cells[0].column_letter].width = (
                length * width + 4
            )

        # set header style ，第一行，标题行，设置颜色和对齐，替换内容
        for cell in ws[1]:
            cell.font = Font(bold=True)
            # borderd
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            # line height
            cell.alignment = cell.alignment.copy(wrap_text=False, vertical="center")
            cell.fill = cell.fill.copy(
                start_color="dddddd", end_color="dddddd", fill_type="solid"
            )
            # set label with config
            for field in self.get_xlsx_fields():
                if field.prop == cell.value:
                    cell.value = field.label
                    break

        # excel 里的特殊类型，不能通过 format string 解决
        if len(self.get_xlsx_fields()) > 0:
            # loop ws data 1~n-1,check type and format
            for row in ws.iter_rows(min_row=1):
                for cell in row:
                    coordinate = cell.coordinate
                    colIndex = coordinate[0]
                    colIndexInt = column_index_from_string(colIndex)
                    # print("format",coordinate,colIndex,colIndexInt)
                    field = self.get_xlsx_fields()[colIndexInt - 1]
                    if field.type == "image":
                        if cell.value:
                            # Image 不支持设置 size、就算修改行高也要再读一遍文件
                            pass
                            # print(cell.value,coordinate,field)
                            # if os.path.exists(str(cell.value)) and os.path.isfile(str(cell.value)):
                            #     img = Image(cell.value)
                            #     cell.value = ""
                            #     ws.add_image(img,coordinate)

        wb.save(file)

    def register(self, router: Router, path=None):
        if not path:
            path = f"/{self.model._meta.model_name}"
        print("register path:", path)
        router.get(f"{path}.detail")(self.get)
        router.get(f"{path}.list")(self.list)
        router.post(f"{path}.create")(self.create)
        router.post(f"{path}.update")(self.update)
        router.post(f"{path}.delete")(self.delete)
        router.get(f"{path}.export")(self.export)
        return self


class ReadOnlyApi(Api):
    def __init__(self, model: models.base.ModelBase, **kwargs):
        super().__init__(
            model,
            config=Api.Config(
                enable_create=False, enable_update=False, enable_delete=False
            ),
            **kwargs,
        )


class PreUserApi(Api):
    forgen_user_field: str

    def __init__(
        self, model: models.base.ModelBase, forgen_user_field="user_id", **kwargs
    ):
        super().__init__(model, **kwargs)
        self.forgen_user_field = forgen_user_field

    def extra_search_condition(self, request: HttpRequest):
        user = request.user
        if not user.is_authenticated:
            raise ApiNotFoundError("Not Authenticated")
        assert isinstance(user, AbstractUser)
        return {self.forgen_user_field: user.pk}

    def register(self, router: Router, path=None):
        if not path:
            path = f"/{self.model._meta.model_name}"
        print("register path:", path)
        router.get(f"{path}.detail")(per_user_api_wrapper(self.get))
        router.get(f"{path}.list")(per_user_api_wrapper(self.list))
        router.post(f"{path}.create")(per_user_api_wrapper(self.create))
        router.put(f"{path}.update")(per_user_api_wrapper(self.update))
        router.delete(f"{path}.delete")(per_user_api_wrapper(self.delete))
        return self


def per_user_api_wrapper(func):
    def wrapper(request: HttpRequest):
        if not request.user.is_authenticated:
            return ApiJsonResponse.error("Not Authenticated", 401)
        return func(request)

    return wrapper
